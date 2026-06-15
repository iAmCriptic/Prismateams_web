import os
from io import BytesIO

from flask import Blueprint, current_app, jsonify, request, send_file, send_from_directory
from openpyxl import Workbook, load_workbook

from app import db
from app.models.assessment import (
    AssessmentCriterion,
    AssessmentList,
    AssessmentListSubject,
    AssessmentRole,
    AssessmentRoom,
    AssessmentStand,
    AssessmentStandType,
    AssessmentUser,
)
from app.utils.assessment_auth import assessment_role_required

excel_uploads_bp = Blueprint("excel_uploads", __name__)

SAMPLE_SPECS = {
    "stands": {
        "filename": "beispiel_staende.xlsx",
        "headers": ["Standname", "Beschreibung", "Raum", "Stand-Typ"],
        "rows": [["Stand Beispiel", "Kurzbeschreibung", "Raum 1", "Essen"]],
    },
    "users": {
        "filename": "beispiel_benutzer.xlsx",
        "headers": ["Benutzername", "Password", "Anzeigename", "Rollen"],
        "rows": [["jury1", "geheim123", "Jury Mitglied", "Bewerter"]],
    },
    "criteria": {
        "filename": "beispiel_kriterien.xlsx",
        "headers": ["Name", "Maximale Punktzahl", "Beschreibung"],
        "rows": [["Kriterium 1", 10, "Beschreibung optional"]],
    },
    "subjects": {
        "filename": "beispiel_bewertungsziele.xlsx",
        "headers": ["Name", "Beschreibung"],
        "rows": [["Maskottchen A", "Optional"]],
    },
}


def _sample_dir():
    return os.path.join(current_app.static_folder, "assessment")


def _build_sample_workbook(which):
    spec = SAMPLE_SPECS[which]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(spec["headers"])
    for row in spec["rows"]:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _row_dict(ws):
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        yield {headers[i]: row[i] for i in range(len(headers))}


def _read_workbook(upload_file, required_columns):
    if not upload_file or not upload_file.filename:
        return None, "Keine Datei empfangen.", 400
    if not upload_file.filename.lower().endswith(".xlsx"):
        return None, "Bitte eine .xlsx-Datei hochladen.", 400
    try:
        workbook = load_workbook(upload_file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return None, f"Datei konnte nicht gelesen werden: {exc}", 400
    ws = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    missing = [col for col in required_columns if col not in headers]
    if missing:
        return None, f"Excel-Datei muss die Spalten enthalten: {', '.join(required_columns)}.", 400
    return ws, None, None


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


@excel_uploads_bp.route("/api/download_sample/<which>")
@assessment_role_required(["Administrator"])
def download_sample(which):
    spec = SAMPLE_SPECS.get(which)
    if not spec:
        return jsonify({"success": False, "message": "Unbekannte Beispieldatei."}), 404
    filename = spec["filename"]
    sample_path = os.path.join(_sample_dir(), filename)
    if os.path.isfile(sample_path):
        return send_from_directory(_sample_dir(), filename, as_attachment=True)
    return send_file(
        _build_sample_workbook(which),
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@excel_uploads_bp.route("/api/import/stands", methods=["POST"])
@assessment_role_required(["Administrator"])
def import_stands():
    ws, error, code = _read_workbook(request.files.get("file"), ["Standname", "Beschreibung", "Raum", "Stand-Typ"])
    if error:
        return jsonify({"success": False, "message": error}), code or 400

    added, updated, errors, rooms_created = 0, 0, [], 0
    default_type = AssessmentStandType.query.order_by(AssessmentStandType.id.asc()).first()

    for index, row in enumerate(_row_dict(ws), start=2):
        name = _clean(row.get("Standname"))
        if not name:
            continue
        description = _clean(row.get("Beschreibung")) or None
        room_name = _clean(row.get("Raum")) or None
        type_name = _clean(row.get("Stand-Typ")) or None

        room = None
        if room_name:
            room = AssessmentRoom.query.filter_by(name=room_name).first()
            if not room:
                room = AssessmentRoom(name=room_name)
                db.session.add(room)
                db.session.flush()
                rooms_created += 1

        stand_type = None
        if type_name:
            stand_type = AssessmentStandType.query.filter_by(name=type_name).first()
            if not stand_type:
                stand_type = AssessmentStandType(name=type_name)
                db.session.add(stand_type)
                db.session.flush()
        elif default_type:
            stand_type = default_type

        existing = AssessmentStand.query.filter_by(name=name).first()
        if existing:
            existing.description = description
            existing.room_id = room.id if room else None
            if stand_type:
                existing.stand_type_id = stand_type.id
            updated += 1
        else:
            db.session.add(
                AssessmentStand(
                    name=name,
                    description=description,
                    room_id=room.id if room else None,
                    stand_type_id=stand_type.id if stand_type else None,
                )
            )
            added += 1

    db.session.commit()
    message = f"Stände importiert: {added} hinzugefügt, {updated} aktualisiert."
    if rooms_created:
        message += f" {rooms_created} Räume automatisch erstellt."
    return jsonify({"success": True, "message": message, "errors": errors})


@excel_uploads_bp.route("/api/import/users", methods=["POST"])
@assessment_role_required(["Administrator"])
def import_users():
    ws, error, code = _read_workbook(
        request.files.get("file"), ["Benutzername", "Password", "Anzeigename", "Rollen"]
    )
    if error:
        return jsonify({"success": False, "message": error}), code or 400

    added, updated, errors = 0, 0, []
    all_roles = {role.name: role for role in AssessmentRole.query.all()}

    for index, row in enumerate(_row_dict(ws), start=2):
        username = _clean(row.get("Benutzername")).lower()
        password = _clean(row.get("Password"))
        display_name = _clean(row.get("Anzeigename")) or username
        roles_text = _clean(row.get("Rollen"))

        if not username or not password or not roles_text:
            errors.append(f"Zeile {index}: Pflichtfelder fehlen.")
            continue

        role_names = [r.strip() for r in roles_text.split(",") if r.strip()]
        roles = [all_roles[name] for name in role_names if name in all_roles]
        if not roles:
            errors.append(f"Zeile {index}: Keine gültigen Rollen für '{username}'.")
            continue

        user = AssessmentUser.query.filter_by(username=username).first()
        if user:
            user.display_name = display_name
            user.set_password(password)
            user.roles = roles
            user.is_admin = any(role.name == "Administrator" for role in roles)
            updated += 1
        else:
            user = AssessmentUser(username=username, display_name=display_name, is_active=True)
            user.set_password(password)
            user.roles = roles
            user.is_admin = any(role.name == "Administrator" for role in roles)
            db.session.add(user)
            added += 1

    db.session.commit()
    message = f"Benutzer importiert: {added} hinzugefügt, {updated} aktualisiert."
    return jsonify({"success": True, "message": message, "errors": errors})


@excel_uploads_bp.route("/api/import/criteria", methods=["POST"])
@assessment_role_required(["Administrator"])
def import_criteria():
    list_id = request.form.get("list_id", type=int) or request.args.get("list_id", type=int)
    if not list_id:
        default_list = AssessmentList.query.filter_by(slug="hauptbewertung").first()
        list_id = default_list.id if default_list else None
    if not list_id:
        return jsonify({"success": False, "message": "Bewertungsliste (list_id) ist erforderlich."}), 400

    ws, error, code = _read_workbook(
        request.files.get("file"), ["Name", "Maximale Punktzahl", "Beschreibung"]
    )
    if error:
        return jsonify({"success": False, "message": error}), code or 400

    added, updated, errors = 0, 0, []

    for index, row in enumerate(_row_dict(ws), start=2):
        name = _clean(row.get("Name"))
        max_raw = row.get("Maximale Punktzahl")
        description = _clean(row.get("Beschreibung")) or None
        if not name or max_raw is None:
            errors.append(f"Zeile {index}: Pflichtfelder fehlen.")
            continue
        try:
            max_score = int(max_raw)
        except (TypeError, ValueError):
            errors.append(f"Zeile {index}: Maximalpunktzahl ist nicht numerisch.")
            continue
        if max_score <= 0:
            errors.append(f"Zeile {index}: Maximalpunktzahl muss > 0 sein.")
            continue

        existing = AssessmentCriterion.query.filter_by(list_id=list_id, name=name).first()
        if existing:
            existing.max_score = max_score
            existing.description = description
            updated += 1
        else:
            db.session.add(
                AssessmentCriterion(
                    list_id=list_id,
                    name=name,
                    max_score=max_score,
                    description=description,
                )
            )
            added += 1

    db.session.commit()
    message = f"Kriterien importiert: {added} hinzugefügt, {updated} aktualisiert."
    return jsonify({"success": True, "message": message, "errors": errors})


@excel_uploads_bp.route("/api/import/subjects", methods=["POST"])
@assessment_role_required(["Administrator"])
def import_subjects():
    list_id = request.form.get("list_id", type=int) or request.args.get("list_id", type=int)
    evaluation_list = AssessmentList.query.get(list_id) if list_id else None
    if not evaluation_list or evaluation_list.subject_mode != "custom":
        return jsonify({"success": False, "message": "Gültige Custom-Bewertungsliste erforderlich."}), 400

    ws, error, code = _read_workbook(request.files.get("file"), ["Name", "Beschreibung"])
    if error:
        return jsonify({"success": False, "message": error}), code or 400

    added, updated, errors = 0, 0, []
    for index, row in enumerate(_row_dict(ws), start=2):
        name = _clean(row.get("Name"))
        if not name:
            continue
        description = _clean(row.get("Beschreibung")) or None
        existing = AssessmentListSubject.query.filter_by(list_id=list_id, name=name).first()
        if existing:
            existing.description = description
            updated += 1
        else:
            db.session.add(
                AssessmentListSubject(list_id=list_id, name=name, description=description)
            )
            added += 1

    db.session.commit()
    message = f"Bewertungsziele importiert: {added} hinzugefügt, {updated} aktualisiert."
    return jsonify({"success": True, "message": message, "errors": errors})
